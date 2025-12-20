import os, sqlite3, io, textwrap, requests
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from campaign_calculator import calculate_campaign_price_and_reach, STATION_DATA, TIME_SLOTS_DATA

app = Flask(__name__)
CORS(app)

TELEGRAM_BOT_TOKEN = '7368212837:AAHqVeOYeIHpJyDXltk-b6eGMmhwdUcM45g'
ADMIN_TELEGRAM_ID = 174046571
DB_PATH = "/tmp/campaigns.db"

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""CREATE TABLE IF NOT EXISTS campaigns (
        id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, campaign_number TEXT,
        radio_stations TEXT, start_date TEXT, end_date TEXT, campaign_days INTEGER,
        time_slots TEXT, campaign_text TEXT, production_option TEXT,
        contact_name TEXT, company TEXT, phone TEXT, email TEXT,
        duration INTEGER, final_price INTEGER, actual_reach INTEGER, ots INTEGER)""")
    conn.commit(); conn.close()

def create_excel_report(row):
    wb = Workbook(); ws = wb.active; ws.title = "Медиаплан"
    blue_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    
    ws.merge_cells("A1:C1")
    ws["A1"] = f"МЕДИАПЛАН КАМПАНИИ #{row['campaign_number']}"
    ws["A1"].fill = blue_fill; ws["A1"].font = white_font; ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:C2"); ws["A2"] = "РАДИО ЗАПАДНОЙ СИБИРИ | ТЮМЕНЬ"; ws["A2"].alignment = Alignment(horizontal="center")
    ws["A4"] = "✅ Ваша заявка принята! Спасибо за доверие!"; ws["A4"].font = bold_font

    ws["A6"] = "📊 ПАРАМЕТРЫ КАМПАНИИ:"; ws["A6"].font = bold_font
    ws.append([f"• Радиостанции: {row['radio_stations']}"])
    ws.append([f"• Период: {row['start_date']} - {row['end_date']} ({row['campaign_days']} дней)"])
    ws.append([f"• Выходов в день: {len(row['time_slots'].split(',')) * len(row['radio_stations'].split(','))}"])
    ws.append([f"• Хронометраж: {row['duration']} сек"])
    
    ws.append([]); ws.append(["📻 ВЫБРАННЫЕ РАДИОСТАНЦИИ:"]); ws.cell(ws.max_row, 1).font = bold_font
    for r in row['radio_stations'].split(','):
        ws.append([f"• {r}: ~{STATION_DATA[r]['reach']*1000:,} слушателей"])

    ws.append([]); ws.append(["🕒 ВЫБРАННЫЕ ВРЕМЕННЫЕ СЛОТЫ:"]); ws.cell(ws.max_row, 1).font = bold_font
    for s_idx in map(int, row['time_slots'].split(',')):
        s = TIME_SLOTS_DATA[s_idx]
        ws.append([f"• {s['time']} - {s['label']}"])

    ws.append([]); ws.append(["🎯 РАСЧЕТНЫЙ ОХВАТ:"]); ws.cell(ws.max_row, 1).font = bold_font
    ws.append([f"• Ежедневный охват: ~{int(row['actual_reach']*0.7):,} чел."])
    ws.append([f"• Общий охват за период: ~{row['actual_reach']:,} чел."])
    ws.append([f"• Рекламных контактов (OTS): {row.get('ots', 0):,}"])

    ws.append([]); ws.append(["💰 ФИНАНСОВАЯ ИНФОРМАЦИЯ:"]); ws.cell(ws.max_row, 1).font = bold_font
    ws.append(["Позиция", "Сумма (₽)"])
    ws.append(["ИТОГО", row['final_price']]); ws.cell(ws.max_row, 1).font = bold_font

    ws.append([]); ws.append(["👤 ВАШИ КОНТАКТЫ:"]); ws.cell(ws.max_row, 1).font = bold_font
    ws.append([f"• Имя: {row['contact_name']}"]); ws.append([f"• Телефон: {row['phone']}"]); ws.append([f"• Компания: {row['company']}"])
    
    ws.append([]); ws.append(["📞 НАШИ КОНТАКТЫ:"]); ws.cell(ws.max_row, 1).font = bold_font
    ws.append(["• Email: alexandra@rzs.ru"]); ws.append(["• Менеджер: Александра Васильева"])

    for col in ['A', 'B']: ws.column_dimensions[col].width = 45
    out = io.BytesIO(); wb.save(out); out.seek(0); return out

@app.route('/')
def index(): return send_from_directory('frontend', 'index.html')

@app.route('/<path:path>')
def static_files(path): return send_from_directory('frontend', path)

@app.route('/api/calculate', methods=['POST'])
def calc():
    res = calculate_campaign_price_and_reach(request.json)
    return jsonify({"success":True, "calculation": {
        "final_price": res[2], "total_reach": res[3], "daily_coverage": res[4], "spots": res[5], "ots": res[6], "bonus": res[1]
    }})

@app.route('/api/create-campaign', methods=['POST'])
def create():
    init_db(); d = request.json; c_num = f"R-{datetime.now().strftime('%H%M%S')}"
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""INSERT INTO campaigns (user_id, campaign_number, radio_stations, start_date, end_date, 
        campaign_days, time_slots, campaign_text, production_option, contact_name, company, phone, email, 
        duration, final_price, actual_reach, ots) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (d.get('user_id'), c_num, ",".join(d.get('selected_radios')), d.get('start_date'), d.get('end_date'), 
         d.get('campaign_days'), ",".join(map(str, d.get('selected_time_slots'))), d.get('campaign_text'), 
         d.get('production_option'), d.get('contact_name'), d.get('company'), d.get('phone'), d.get('email'), 
         d.get('duration'), d.get('final_price'), d.get('total_reach'), d.get('ots')))
    conn.commit(); conn.close()
    
    row_dict = d.copy(); row_dict['campaign_number'] = c_num
    row_dict['time_slots'] = ",".join(map(str, d.get('selected_time_slots')))
    row_dict['radio_stations'] = ",".join(d.get('selected_radios'))
    excel = create_excel_report(row_dict)
    requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument", 
        files={'document': (f'Mediaplan_{c_num}.xlsx', excel)}, 
        data={'chat_id': d.get('user_id'), 'caption': f'Ваш медиаплан {c_num} готов!'})
    
    return jsonify({"success":True, "campaign_number": c_num})

@app.route('/api/user-campaigns/<int:user_id>')
def history(user_id):
    init_db(); conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM campaigns WHERE user_id = ? ORDER BY id DESC", (user_id,)).fetchall()
    return jsonify({"success":True, "campaigns": [dict(r) for r in rows]})

@app.route('/api/confirmation/<num>')
def conf(num):
    init_db(); conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM campaigns WHERE campaign_number = ?", (num,)).fetchone()
    return jsonify({"success":True, "campaign": dict(row)}) if row else jsonify({"success":False})

if __name__ == '__main__': init_db(); app.run(host='0.0.0.0', port=5000)
